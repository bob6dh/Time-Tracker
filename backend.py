import csv
import glob
import json
import os
import shutil
import sys
import time
import calendar
import uuid
from datetime import datetime, date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from PySide6.QtCore import (
    QObject, Property, Signal, Slot, QTimer, QAbstractListModel,
    Qt, QModelIndex, QByteArray, QUrl,
)

if getattr(sys, "frozen", False):
    _app_dir = os.path.dirname(sys.executable)
else:
    _app_dir = os.path.dirname(os.path.abspath(__file__))

DATA_FILE = os.path.join(_app_dir, "tracker_data.json")
BAK_FILE = DATA_FILE + ".bak"
BACKUP_DIR = os.path.join(_app_dir, "backups")
BACKUP_RETENTION_DAYS = 30


def _url_to_path(url_str: str) -> str:
    """Convert a file:// URL from QML FileDialog to a local filesystem path."""
    path = QUrl(str(url_str)).toLocalFile()
    return path if path else str(url_str)


def _proj_name(p):
    """Return the name string from either a legacy string project or a dict project."""
    return p["name"] if isinstance(p, dict) else p


def _is_valid_data(data):
    return isinstance(data, dict) and "dailyLogs" in data and "projects" in data


def _try_load(path):
    if not path or not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return None
    return data if _is_valid_data(data) else None


def _daily_backups():
    """Return (date, path) pairs for dated snapshots in BACKUP_DIR, newest first.

    Excludes pre-import snapshots, whose stems don't parse as a plain ISO date.
    """
    if not os.path.isdir(BACKUP_DIR):
        return []
    found = []
    for path in glob.glob(os.path.join(BACKUP_DIR, "tracker_data_*.json")):
        stem = os.path.basename(path)[len("tracker_data_"):-len(".json")]
        try:
            found.append((date.fromisoformat(stem), path))
        except ValueError:
            continue
    return sorted(found, reverse=True)


def load_data():
    # A corrupted/missing main file falls back to the rolling backup, then the
    # most recent daily snapshot, before ever resetting to an empty dataset —
    # otherwise the next save would permanently overwrite recoverable data.
    data = _try_load(DATA_FILE)
    if data is None:
        data = _try_load(BAK_FILE)
    if data is None:
        backups = _daily_backups()
        if backups:
            data = _try_load(backups[0][1])
    if data is None:
        data = {"projects": [], "checkInInterval": 30, "dailyLogs": {}}
    # Migrate legacy string-format projects to dict format
    data["projects"] = [
        p if isinstance(p, dict)
        else {"name": p, "billingCode": "", "billable": True}
        for p in data.get("projects", [])
    ]
    data.setdefault("tasks", [])
    for t in data["tasks"]:
        t.setdefault("dueDate", None)
    return data


def _prune_old_backups():
    cutoff = date.today() - timedelta(days=BACKUP_RETENTION_DAYS)
    for snap_date, path in _daily_backups():
        if snap_date < cutoff:
            try:
                os.remove(path)
            except OSError:
                pass


def _copy_into_backup_dir(filename):
    """Best-effort copy of the current DATA_FILE into BACKUP_DIR under the given
    name. Returns True if a copy was written."""
    if not os.path.exists(DATA_FILE):
        return False
    os.makedirs(BACKUP_DIR, exist_ok=True)
    try:
        shutil.copyfile(DATA_FILE, os.path.join(BACKUP_DIR, filename))
        return True
    except OSError:
        return False


def _snapshot_daily_backup():
    """Copy today's pre-save state into BACKUP_DIR, once per calendar day."""
    filename = f"tracker_data_{date.today().isoformat()}.json"
    if os.path.exists(os.path.join(BACKUP_DIR, filename)):
        return
    if _copy_into_backup_dir(filename):
        _prune_old_backups()


def snapshot_before_overwrite(reason):
    """Explicit undo point before some action wholesale-replaces all in-memory
    data (import, restore-from-backup). `reason` (e.g. "pre-import",
    "pre-restore") becomes part of the backup filename."""
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    _copy_into_backup_dir(f"tracker_data_{reason}_{stamp}.json")


def _named_backups(reason):
    """Return (datetime, path) pairs for tracker_data_{reason}_<timestamp>.json
    snapshots written by snapshot_before_overwrite(reason), newest first."""
    if not os.path.isdir(BACKUP_DIR):
        return []
    prefix = f"tracker_data_{reason}_"
    found = []
    for path in glob.glob(os.path.join(BACKUP_DIR, f"{prefix}*.json")):
        stem = os.path.basename(path)[len(prefix):-len(".json")]
        try:
            found.append((datetime.strptime(stem, "%Y-%m-%d_%H%M%S"), path))
        except ValueError:
            continue
    return sorted(found, reverse=True)


def _replace_with_retry(src, dst, attempts=5, delay=0.05):
    """os.replace() can transiently raise WinError 5 (Access is denied) when dst
    lives in a cloud-synced folder (OneDrive, Dropbox, ...) that briefly locks the
    file for scanning/upload right after a write. Retry with backoff before
    falling back to a plain (non-atomic) copy so a save is never lost.

    This runs synchronously on the Qt main thread (called from save_data(), which
    every mutating Slot uses), so the backoff is kept short — worst case ~0.75s
    total across 5 attempts — to avoid noticeably freezing the UI."""
    last_err = None
    for i in range(attempts):
        try:
            os.replace(src, dst)
            return
        except PermissionError as e:
            last_err = e
            time.sleep(delay * (i + 1))
    try:
        shutil.copyfile(src, dst)
        os.remove(src)
    except OSError:
        raise last_err


def save_data(data):
    if os.path.exists(DATA_FILE):
        try:
            shutil.copyfile(DATA_FILE, BAK_FILE)
        except OSError:
            pass
        _snapshot_daily_backup()
    tmp_path = DATA_FILE + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    _replace_with_retry(tmp_path, DATA_FILE)  # atomic on Windows and POSIX, with OneDrive-lock retry


def fmt_time(seconds):
    seconds = int(seconds)
    h, m, s = seconds // 3600, (seconds % 3600) // 60, seconds % 60
    if h > 0:
        return f"{h}h {m}m"
    if m > 0:
        return f"{m}m {s}s"
    return f"{s}s"


def fmt_date(d):
    try:
        return datetime.strptime(d, "%Y-%m-%d").strftime("%a, %b %d, %Y")
    except Exception:
        return d


# ── Project list model ──────────────────────────────────────────


class ProjectListModel(QAbstractListModel):
    NameRole = Qt.UserRole + 1
    TodayTimeRole = Qt.UserRole + 2
    IsActiveRole = Qt.UserRole + 3
    BillingCodeRole = Qt.UserRole + 4
    BillableRole = Qt.UserRole + 5

    def __init__(self, backend, parent=None):
        super().__init__(parent)
        self._backend = backend

    def roleNames(self):
        return {
            self.NameRole: QByteArray(b"name"),
            self.TodayTimeRole: QByteArray(b"todayTime"),
            self.IsActiveRole: QByteArray(b"isActive"),
            self.BillingCodeRole: QByteArray(b"billingCode"),
            self.BillableRole: QByteArray(b"billable"),
        }

    def _visible(self):
        """Return only non-archived projects for the timer UI.
        Archived projects are intentionally NOT filtered from dailyLogs,
        so their data remains fully visible in history, reports, and exports.
        """
        return [
            p for p in self._backend._data["projects"]
            if not (isinstance(p, dict) and p.get("archived", False))
        ]

    def rowCount(self, parent=QModelIndex()):
        return len(self._visible())

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        proj = self._visible()[index.row()]
        name = _proj_name(proj)
        if role == self.NameRole:
            return name
        if role == self.TodayTimeRole:
            return fmt_time(self._backend._get_today_total(name))
        if role == self.IsActiveRole:
            return self._backend._active_project == name
        if role == self.BillingCodeRole:
            return proj.get("billingCode", "") if isinstance(proj, dict) else ""
        if role == self.BillableRole:
            return proj.get("billable", True) if isinstance(proj, dict) else True
        return None

    def refresh(self):
        self.beginResetModel()
        self.endResetModel()


# ── Task list model ──────────────────────────────────────────────


class TaskListModel(QAbstractListModel):
    IdRole = Qt.UserRole + 1
    TitleRole = Qt.UserRole + 2
    ProjectRole = Qt.UserRole + 3
    IsActiveRole = Qt.UserRole + 4
    TimeRole = Qt.UserRole + 5
    DueDateRole = Qt.UserRole + 6
    DueDateLabelRole = Qt.UserRole + 7

    def __init__(self, backend, parent=None):
        super().__init__(parent)
        self._backend = backend

    def roleNames(self):
        return {
            self.IdRole: QByteArray(b"taskId"),
            self.TitleRole: QByteArray(b"title"),
            self.ProjectRole: QByteArray(b"project"),
            self.IsActiveRole: QByteArray(b"isActive"),
            self.TimeRole: QByteArray(b"timeText"),
            self.DueDateRole: QByteArray(b"dueDate"),
            self.DueDateLabelRole: QByteArray(b"dueDateLabel"),
        }

    def _visible(self):
        """Return pending (not-done) tasks for the Tasks tab, sorted by due date
        (undated tasks last), then by creation date."""
        tasks = [t for t in self._backend._data.get("tasks", []) if not t.get("done")]
        return sorted(
            tasks,
            key=lambda t: (t.get("dueDate") or "9999-12-31", t.get("createdDate") or ""),
        )

    def rowCount(self, parent=QModelIndex()):
        return len(self._visible())

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        task = self._visible()[index.row()]
        if role == self.IdRole:
            return task["id"]
        if role == self.TitleRole:
            return task["title"]
        if role == self.ProjectRole:
            return task["project"]
        if role == self.IsActiveRole:
            return self._backend._active_task_id == task["id"]
        if role == self.TimeRole:
            return fmt_time(self._backend._task_seconds(task["id"]))
        if role == self.DueDateRole:
            return task.get("dueDate") or ""
        if role == self.DueDateLabelRole:
            due = task.get("dueDate")
            return fmt_date(due) if due else ""
        return None

    def refresh(self):
        self.beginResetModel()
        self.endResetModel()

    def refresh_active_time(self):
        """Update just the active task's row (its live elapsed time), without
        resetting the whole list — used on every timer tick so the Tasks tab
        doesn't re-scan and re-render every pending row once a second."""
        active_id = self._backend._active_task_id
        if not active_id:
            return
        tasks = self._visible()
        for row, task in enumerate(tasks):
            if task["id"] == active_id:
                idx = self.index(row)
                self.dataChanged.emit(idx, idx, [self.TimeRole])
                return


# ── History list model ──────────────────────────────────────────


class HistoryListModel(QAbstractListModel):
    DateKeyRole = Qt.UserRole + 1
    DateLabelRole = Qt.UserRole + 2
    ProjectCountRole = Qt.UserRole + 3
    TotalTimeRole = Qt.UserRole + 4

    def __init__(self, backend, parent=None):
        super().__init__(parent)
        self._backend = backend

    def roleNames(self):
        return {
            self.DateKeyRole: QByteArray(b"dateKey"),
            self.DateLabelRole: QByteArray(b"dateLabel"),
            self.ProjectCountRole: QByteArray(b"projectCount"),
            self.TotalTimeRole: QByteArray(b"totalTime"),
        }

    def rowCount(self, parent=QModelIndex()):
        return len(self._sorted_days())

    def _sorted_days(self):
        days = set(self._backend._data["dailyLogs"].keys())
        # Include today if there's an active session even before first stop
        if self._backend._active_project and self._backend._session_start:
            days.add(date.today().isoformat())
        return sorted(days, reverse=True)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        day = self._sorted_days()[index.row()]
        log = self._backend._data["dailyLogs"].get(day, {})
        if role == self.DateKeyRole:
            return day
        if role == self.DateLabelRole:
            return fmt_date(day)
        if role == self.ProjectCountRole:
            projects = set(log.keys())
            today_str = date.today().isoformat()
            if day == today_str and self._backend._active_project:
                projects.add(self._backend._active_project)
            n = len(projects)
            return f"{n} project{'s' if n != 1 else ''}"
        if role == self.TotalTimeRole:
            total = sum(v["seconds"] for v in log.values())
            today_str = date.today().isoformat()
            if day == today_str and self._backend._active_project and self._backend._session_start:
                total += int(time.time() - self._backend._session_start)
            return fmt_time(total)
        return None

    def refresh(self):
        self.beginResetModel()
        self.endResetModel()


# ── Day detail list model ───────────────────────────────────────


class DayDetailModel(QAbstractListModel):
    ProjectRole = Qt.UserRole + 1
    TimeRole = Qt.UserRole + 2
    DescriptionRole = Qt.UserRole + 3

    def __init__(self, backend, parent=None):
        super().__init__(parent)
        self._backend = backend
        self._items = []

    def roleNames(self):
        return {
            self.ProjectRole: QByteArray(b"project"),
            self.TimeRole: QByteArray(b"time"),
            self.DescriptionRole: QByteArray(b"description"),
        }

    def rowCount(self, parent=QModelIndex()):
        return len(self._items)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        item = self._items[index.row()]
        if role == self.ProjectRole:
            return item["project"]
        if role == self.TimeRole:
            return item["time"]
        if role == self.DescriptionRole:
            return item["description"]
        return None

    def load_day(self, day_key):
        self.beginResetModel()
        log = self._backend._data["dailyLogs"].get(day_key, {})
        self._items = [
            {
                "project": proj,
                "time": fmt_time(info["seconds"]),
                "description": info.get("description", ""),
            }
            for proj, info in log.items()
        ]
        self.endResetModel()


# ── EOD model ───────────────────────────────────────────────────


class EodModel(QAbstractListModel):
    ProjectRole = Qt.UserRole + 1
    DescriptionRole = Qt.UserRole + 2
    TimeRole = Qt.UserRole + 3

    def __init__(self, backend, parent=None):
        super().__init__(parent)
        self._backend = backend
        self._items = []

    def roleNames(self):
        return {
            self.ProjectRole: QByteArray(b"project"),
            self.DescriptionRole: QByteArray(b"description"),
            self.TimeRole: QByteArray(b"timeText"),
        }

    def rowCount(self, parent=QModelIndex()):
        return len(self._items)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        item = self._items[index.row()]
        if role == self.ProjectRole:
            return item["project"]
        if role == self.DescriptionRole:
            return item["description"]
        if role == self.TimeRole:
            return item["timeText"]
        return None

    @Slot()
    def load(self):
        self.beginResetModel()
        today = date.today().isoformat()
        log = self._backend._data["dailyLogs"].get(today, {})
        self._items = [
            {
                "project": proj,
                "description": info.get("description", ""),
                "timeText": fmt_time(info.get("seconds", 0)),
            }
            for proj, info in log.items()
        ]
        self.endResetModel()

    @Slot(int, str)
    def setDescription(self, index, desc):
        if 0 <= index < len(self._items):
            self._items[index]["description"] = desc


# ── Report model ───────────────────────────────────────────────


class ReportModel(QAbstractListModel):
    ProjectRole = Qt.UserRole + 1
    TimeRole = Qt.UserRole + 2
    SecondsRole = Qt.UserRole + 3

    def __init__(self, backend, parent=None):
        super().__init__(parent)
        self._backend = backend
        self._items = []

    def roleNames(self):
        return {
            self.ProjectRole: QByteArray(b"project"),
            self.TimeRole: QByteArray(b"time"),
            self.SecondsRole: QByteArray(b"seconds"),
        }

    def rowCount(self, parent=QModelIndex()):
        return len(self._items)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        item = self._items[index.row()]
        if role == self.ProjectRole:
            return item["project"]
        if role == self.TimeRole:
            return fmt_time(item["seconds"])
        if role == self.SecondsRole:
            return item["seconds"]
        return None

    def load(self, date_keys):
        totals = {}
        for dk in date_keys:
            log = self._backend._data["dailyLogs"].get(dk, {})
            for proj, info in log.items():
                totals[proj] = totals.get(proj, 0) + info["seconds"]
        # Include active session time if today is in the range
        today = date.today().isoformat()
        if today in date_keys and self._backend._active_project and self._backend._session_start:
            proj = self._backend._active_project
            totals[proj] = totals.get(proj, 0) + int(time.time() - self._backend._session_start)
        self.beginResetModel()
        self._items = sorted(
            [{"project": p, "seconds": s} for p, s in totals.items()],
            key=lambda x: x["seconds"], reverse=True,
        )
        self.endResetModel()

    @property
    def total_seconds(self):
        return sum(i["seconds"] for i in self._items)


# ── Main backend ────────────────────────────────────────────────


class TimeTrackerBackend(QObject):
    # Signals
    activeProjectChanged = Signal()
    elapsedChanged = Signal()
    elapsedTextChanged = Signal()
    checkInIntervalChanged = Signal()
    showCheckIn = Signal()
    showEod = Signal()
    hasTodayLogsChanged = Signal()
    reportPeriodChanged = Signal()
    reportLabelChanged = Signal()
    reportTotalChanged = Signal()
    reportTotalSecondsChanged = Signal()
    exportDone = Signal(str, bool)  # (message, success)
    summaryChanged = Signal()
    jsonTransferDone = Signal(str, bool)  # (message, success)
    archivedProjectsChanged = Signal()
    completedTasksChanged = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._data = load_data()
        self._active_project = None
        self._active_task_id = None
        self._session_start = None
        self._elapsed = 0
        self._last_checkin = None
        self._checkin_shown_at = None
        self._eod_dismissed = False

        self.INACTIVITY_TIMEOUT = 30 * 60  # seconds before auto-stopping if no check-in response

        self._project_model = ProjectListModel(self)
        self._task_model = TaskListModel(self)
        self._history_model = HistoryListModel(self)
        self._day_detail_model = DayDetailModel(self)
        self._eod_model = EodModel(self)
        self._report_model = ReportModel(self)
        self._report_period = "day"
        self._report_offset = 0

        self._timer = QTimer(self)
        self._timer.setInterval(1000)
        self._timer.timeout.connect(self._tick)
        self._timer.start()

        # Restore any active session that was running when the app last closed
        self._restore_active_session()

    # ── Properties ──

    @Property(str, notify=activeProjectChanged)
    def activeProject(self):
        return self._active_project or ""

    @Property(int, notify=elapsedChanged)
    def elapsed(self):
        return self._elapsed

    @Property(str, notify=elapsedTextChanged)
    def elapsedText(self):
        return fmt_time(self._elapsed)

    @Property(int, notify=checkInIntervalChanged)
    def checkInInterval(self):
        return self._data["checkInInterval"]

    @Property(bool, notify=hasTodayLogsChanged)
    def hasTodayLogs(self):
        today = date.today().isoformat()
        return bool(self._data["dailyLogs"].get(today))

    @Property(QObject, constant=True)
    def projectModel(self):
        return self._project_model

    @Property(QObject, constant=True)
    def taskModel(self):
        return self._task_model

    @Property(str, notify=activeProjectChanged)
    def activeTaskTitle(self):
        if not self._active_task_id:
            return ""
        task = self._task_by_id(self._active_task_id)
        return task["title"] if task else ""

    @Property(QObject, constant=True)
    def historyModel(self):
        return self._history_model

    @Property(QObject, constant=True)
    def dayDetailModel(self):
        return self._day_detail_model

    @Property(QObject, constant=True)
    def eodModel(self):
        return self._eod_model

    @Property(QObject, constant=True)
    def reportModel(self):
        return self._report_model

    @Property(str, notify=reportPeriodChanged)
    def reportPeriod(self):
        return self._report_period

    @Property(str, notify=reportLabelChanged)
    def reportLabel(self):
        return self._get_report_label()

    @Property(str, notify=reportTotalChanged)
    def reportTotal(self):
        return fmt_time(self._report_model.total_seconds)

    @Property(int, notify=reportTotalSecondsChanged)
    def reportTotalSeconds(self):
        return self._report_model.total_seconds

    @Property(int, constant=True)
    def inactivityTimeoutSecs(self):
        return self.INACTIVITY_TIMEOUT

    @Property(str, notify=summaryChanged)
    def todayTotal(self):
        today = date.today().isoformat()
        secs = sum(v["seconds"] for v in self._data["dailyLogs"].get(today, {}).values())
        if self._active_project and self._session_start:
            secs += int(time.time() - self._session_start)
        return fmt_time(secs) if secs > 0 else "0m"

    @Property(str, notify=summaryChanged)
    def weekTotal(self):
        today = date.today()
        monday = today - timedelta(days=today.weekday())
        secs = 0
        for i in range(7):
            dk = (monday + timedelta(days=i)).isoformat()
            secs += sum(v["seconds"] for v in self._data["dailyLogs"].get(dk, {}).values())
        if self._active_project and self._session_start:
            secs += int(time.time() - self._session_start)
        return fmt_time(secs) if secs > 0 else "0m"

    @Property(str, notify=summaryChanged)
    def monthTotal(self):
        today = date.today()
        days_in_month = calendar.monthrange(today.year, today.month)[1]
        secs = 0
        for d in range(1, days_in_month + 1):
            dk = date(today.year, today.month, d).isoformat()
            secs += sum(v["seconds"] for v in self._data["dailyLogs"].get(dk, {}).values())
        if self._active_project and self._session_start:
            secs += int(time.time() - self._session_start)
        return fmt_time(secs) if secs > 0 else "0m"

    def _billable_project_names(self) -> set:
        names = set()
        for p in self._data["projects"]:
            if isinstance(p, dict):
                if p.get("billable", True):
                    names.add(p["name"])
            else:
                names.add(str(p))
        return names

    @Property(str, notify=summaryChanged)
    def todayBillable(self):
        billable = self._billable_project_names()
        today = date.today().isoformat()
        secs = sum(
            v["seconds"] for k, v in self._data["dailyLogs"].get(today, {}).items()
            if k in billable
        )
        if self._active_project and self._session_start and self._active_project in billable:
            secs += int(time.time() - self._session_start)
        return fmt_time(secs) if secs > 0 else "0m"

    @Property(str, notify=summaryChanged)
    def weekBillable(self):
        billable = self._billable_project_names()
        today = date.today()
        monday = today - timedelta(days=today.weekday())
        secs = 0
        for i in range(7):
            dk = (monday + timedelta(days=i)).isoformat()
            secs += sum(
                v["seconds"] for k, v in self._data["dailyLogs"].get(dk, {}).items()
                if k in billable
            )
        if self._active_project and self._session_start and self._active_project in billable:
            secs += int(time.time() - self._session_start)
        return fmt_time(secs) if secs > 0 else "0m"

    @Property(str, notify=summaryChanged)
    def monthBillable(self):
        billable = self._billable_project_names()
        today = date.today()
        days_in_month = calendar.monthrange(today.year, today.month)[1]
        secs = 0
        for d in range(1, days_in_month + 1):
            dk = date(today.year, today.month, d).isoformat()
            secs += sum(
                v["seconds"] for k, v in self._data["dailyLogs"].get(dk, {}).items()
                if k in billable
            )
        if self._active_project and self._session_start and self._active_project in billable:
            secs += int(time.time() - self._session_start)
        return fmt_time(secs) if secs > 0 else "0m"

    # ── Slots ──

    @Slot(str, str, bool)
    def addProject(self, name: str, billing_code: str, billable: bool):
        name = name.strip()
        if not name:
            return
        existing = [_proj_name(p) for p in self._data["projects"]]
        if name in existing:
            return
        self._data["projects"].append({
            "name": name,
            "billingCode": billing_code.strip(),
            "billable": billable,
        })
        save_data(self._data)
        self._project_model.refresh()

    @Slot(str)
    def archiveProject(self, name: str):
        if self._active_project == name:
            self.stopTimer()
        for p in self._data["projects"]:
            if _proj_name(p) == name:
                if isinstance(p, dict):
                    p["archived"] = True
                break
        save_data(self._data)
        self._project_model.refresh()
        self.archivedProjectsChanged.emit()

    @Slot(str)
    def reinstateProject(self, name: str):
        for p in self._data["projects"]:
            if _proj_name(p) == name:
                if isinstance(p, dict):
                    p["archived"] = False
                break
        save_data(self._data)
        self._project_model.refresh()
        self.archivedProjectsChanged.emit()

    @Slot(result="QVariantList")
    def getArchivedProjects(self):
        return [
            {
                "name": _proj_name(p),
                "billingCode": p.get("billingCode", "") if isinstance(p, dict) else "",
                "billable": p.get("billable", True) if isinstance(p, dict) else True,
            }
            for p in self._data["projects"]
            if isinstance(p, dict) and p.get("archived", False)
        ]

    @Slot(str)
    def startProject(self, name):
        self._start(name, None)

    @Slot(str)
    def startTask(self, task_id):
        task = self._task_by_id(task_id)
        if not task:
            return
        self._start(task["project"], task_id)

    def _start(self, name, task_id):
        if self._active_project and self._session_start:
            self._log_time(int(time.time() - self._session_start))
        self._active_project = name
        self._active_task_id = task_id
        self._session_start = time.time()
        self._last_checkin = time.time()
        self._elapsed = 0
        self._save_active_session()
        self.activeProjectChanged.emit()
        self.elapsedChanged.emit()
        self.elapsedTextChanged.emit()
        self.summaryChanged.emit()
        self._project_model.refresh()
        self._history_model.refresh()
        self._task_model.refresh()
        self.hasTodayLogsChanged.emit()

    @Slot()
    def stopTimer(self):
        if self._active_project and self._session_start:
            self._log_time(int(time.time() - self._session_start))
        self._active_project = None
        self._active_task_id = None
        self._session_start = None
        self._elapsed = 0
        self._last_checkin = None
        self._checkin_shown_at = None
        self._clear_active_session()
        self.activeProjectChanged.emit()
        self.elapsedChanged.emit()
        self.elapsedTextChanged.emit()
        self.summaryChanged.emit()
        self._project_model.refresh()
        self._history_model.refresh()
        self._task_model.refresh()
        self.hasTodayLogsChanged.emit()

    # ── Tasks ──

    @Slot(str, str, str)
    def addTask(self, title: str, project: str, due_date: str):
        title = title.strip()
        if not title or not project:
            return
        self._data.setdefault("tasks", []).append({
            "id": uuid.uuid4().hex,
            "title": title,
            "project": project,
            "done": False,
            "createdDate": date.today().isoformat(),
            "completedDate": None,
            "dueDate": due_date or None,
        })
        save_data(self._data)
        self._task_model.refresh()

    @Slot(str, str, str, str)
    def editTask(self, task_id: str, title: str, project: str, due_date: str):
        title = title.strip()
        if not title or not project:
            return
        task = self._task_by_id(task_id)
        if not task:
            return
        task["title"] = title
        task["project"] = project
        task["dueDate"] = due_date or None
        save_data(self._data)
        self._task_model.refresh()

    @Slot(str)
    def completeTask(self, task_id: str):
        if self._active_task_id == task_id:
            self.stopTimer()
        task = self._task_by_id(task_id)
        if not task:
            return
        task["done"] = True
        task["completedDate"] = date.today().isoformat()
        save_data(self._data)
        self._task_model.refresh()
        self.completedTasksChanged.emit()

    @Slot(str)
    def reopenTask(self, task_id: str):
        task = self._task_by_id(task_id)
        if not task:
            return
        task["done"] = False
        task["completedDate"] = None
        save_data(self._data)
        self._task_model.refresh()
        self.completedTasksChanged.emit()

    @Slot(str)
    def deleteTask(self, task_id: str):
        if self._active_task_id == task_id:
            self.stopTimer()
        self._data["tasks"] = [
            t for t in self._data.get("tasks", []) if t["id"] != task_id
        ]
        save_data(self._data)
        self._task_model.refresh()
        self.completedTasksChanged.emit()

    @Slot(result="QVariantList")
    def getCompletedTasks(self):
        return [
            {
                "id": t["id"],
                "title": t["title"],
                "project": t["project"],
                "time": fmt_time(self._task_seconds(t["id"])),
            }
            for t in self._data.get("tasks", [])
            if t.get("done")
        ]

    @Slot(result="QVariantList")
    def getProjectNames(self):
        return [
            _proj_name(p) for p in self._data["projects"]
            if not (isinstance(p, dict) and p.get("archived", False))
        ]

    @Slot(int)
    def setCheckInInterval(self, minutes):
        self._data["checkInInterval"] = minutes
        save_data(self._data)
        self.checkInIntervalChanged.emit()

    @Slot()
    def clearHistory(self):
        self._data["dailyLogs"] = {}
        save_data(self._data)
        self._history_model.refresh()
        self.hasTodayLogsChanged.emit()

    @Slot(str)
    def openDayDetail(self, day_key):
        self._day_detail_model.load_day(day_key)

    @Slot(str, result=str)
    def dayDetailTitle(self, day_key):
        return fmt_date(day_key)

    @Slot()
    def checkInYes(self):
        self._last_checkin = time.time()
        self._checkin_shown_at = None

    @Slot()
    def checkInNo(self):
        self.stopTimer()

    @Slot()
    def openEodDialog(self):
        self._eod_model.load()
        self.showEod.emit()

    @Slot()
    def saveEod(self):
        today = date.today().isoformat()
        for item in self._eod_model._items:
            proj = item["project"]
            desc = item["description"]
            if today in self._data["dailyLogs"] and proj in self._data["dailyLogs"][today]:
                self._data["dailyLogs"][today][proj]["description"] = desc
        save_data(self._data)
        self._eod_dismissed = True

    @Slot()
    def dismissEod(self):
        self._eod_dismissed = True

    @Slot(result="QVariantList")
    def getDatesWithData(self):
        return list(self._data["dailyLogs"].keys())

    @Slot(str, str)
    def removeProjectFromDay(self, day_key: str, project_name: str):
        logs = self._data["dailyLogs"]
        if day_key in logs and project_name in logs[day_key]:
            del logs[day_key][project_name]
            if not logs[day_key]:          # remove empty day entry
                del logs[day_key]
            save_data(self._data)
            self._history_model.refresh()
            self._day_detail_model.load_day(day_key)
            if day_key == date.today().isoformat():
                self._project_model.refresh()
                self.hasTodayLogsChanged.emit()
            self.summaryChanged.emit()

    @Slot(str, str)
    def addProjectToDay(self, day_key: str, project_name: str):
        logs = self._data["dailyLogs"]
        if day_key not in logs:
            logs[day_key] = {}
        if project_name not in logs[day_key]:
            logs[day_key][project_name] = {
                "seconds": 0,
                "sessions": [],
                "description": "",
            }
            save_data(self._data)
            self._history_model.refresh()
            self._day_detail_model.load_day(day_key)
            if day_key == date.today().isoformat():
                self._project_model.refresh()
                self.hasTodayLogsChanged.emit()
            self.summaryChanged.emit()

    @Slot(str, result="QVariantList")
    def getDayData(self, day_key: str):
        log = self._data["dailyLogs"].get(day_key, {})
        return [
            {
                "project": proj,
                "seconds": info["seconds"],
                "sessions": info.get("sessions", []),
                "description": info.get("description", ""),
            }
            for proj, info in log.items()
        ]

    @Slot(str, str, str)
    def saveDaySessions(self, day_key: str, project: str, sessions_json: str):
        import json as _json
        sessions = _json.loads(sessions_json)
        logs = self._data["dailyLogs"]
        if day_key not in logs or project not in logs[day_key]:
            return
        logs[day_key][project]["sessions"] = sessions
        logs[day_key][project]["seconds"] = sum(
            (s["end"] - s["start"]) * 60 for s in sessions
        )
        save_data(self._data)
        self._history_model.refresh()
        self._day_detail_model.load_day(day_key)
        self._task_model.refresh()
        if day_key == date.today().isoformat():
            self._project_model.refresh()
            self.hasTodayLogsChanged.emit()

    @Slot(str, str, str)
    def saveProjectDescription(self, day_key: str, project: str, description: str):
        logs = self._data["dailyLogs"]
        if day_key not in logs or project not in logs[day_key]:
            return
        logs[day_key][project]["description"] = description
        save_data(self._data)
        self._day_detail_model.load_day(day_key)

    @Slot()
    def refreshModels(self):
        self._project_model.refresh()
        self._history_model.refresh()
        self._task_model.refresh()
        self.hasTodayLogsChanged.emit()

    @Slot(str)
    def setReportPeriod(self, period):
        self._report_period = period
        self._report_offset = 0
        self._load_report()
        self.reportPeriodChanged.emit()

    @Slot()
    def reportPrev(self):
        self._report_offset -= 1
        self._load_report()

    @Slot()
    def reportNext(self):
        if self._report_offset < 0:
            self._report_offset += 1
            self._load_report()

    @Slot()
    def refreshReport(self):
        self._load_report()

    def _collect_monthly_rows(self, year, month):
        """Return (month_name, rows, total_hours) for a given year/month, where
        rows is a list of (date_label, project, hours, description) tuples —
        the shared data both the Excel and CSV monthly exports render."""
        month_name = datetime(year, month, 1).strftime("%B %Y")
        days_in_month = calendar.monthrange(year, month)[1]
        rows = []
        total_hours = 0.0
        for d in range(1, days_in_month + 1):
            day_key = f"{year}-{month:02d}-{d:02d}"
            log = self._data["dailyLogs"].get(day_key, {})
            for proj, info in sorted(log.items()):
                secs = info.get("seconds", 0)
                # Round to nearest 10 minutes, express as decimal hours
                hours = round(secs / 600) * 10 / 60
                desc = info.get("description", "")
                date_label = datetime(year, month, d).strftime("%a, %b %d")
                rows.append((date_label, proj, hours, desc))
                total_hours += hours
        return month_name, rows, total_hours

    @Slot(str, str)
    def exportMonthlyReport(self, year_month, file_path):
        """Export a monthly report to an Excel file.

        Args:
            year_month: "YYYY-MM" string
            file_path:  Absolute path (may start with "file://")
        """
        try:
            file_path = _url_to_path(file_path)
            if not file_path.endswith(".xlsx"):
                file_path += ".xlsx"

            year, month = int(year_month[:4]), int(year_month[5:7])
            month_name, rows, total_hours = self._collect_monthly_rows(year, month)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Report"

            # ── Styles ───────────────────────────────────────────────
            dark = "1F2937"
            blue = "2563EB"
            header_fill = PatternFill("solid", fgColor=dark)
            alt_fill    = PatternFill("solid", fgColor="F9FAFB")
            total_fill  = PatternFill("solid", fgColor="EEF4FF")

            header_font = Font(bold=True, color="FFFFFF", size=11)
            title_font  = Font(bold=True, color=dark, size=14)
            total_font  = Font(bold=True, color=dark, size=11)
            total_lbl_font = Font(bold=True, color=blue, size=11)

            thin = Side(style="thin", color="E5E7EB")
            cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

            center = Alignment(horizontal="center", vertical="center")
            left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
            right  = Alignment(horizontal="right",  vertical="center")

            # ── Title row ────────────────────────────────────────────
            ws.merge_cells("A1:D1")
            title_cell = ws["A1"]
            title_cell.value = f"Monthly Report — {month_name}"
            title_cell.font = title_font
            title_cell.alignment = left
            ws.row_dimensions[1].height = 28

            ws.append([])  # blank row 2

            # ── Header row ───────────────────────────────────────────
            headers = ["Date", "Project", "Hours", "Description"]
            ws.append(headers)
            for col, _ in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = cell_border
            ws.row_dimensions[3].height = 22

            # ── Data rows ────────────────────────────────────────────
            for i, (date_lbl, proj, hrs, desc) in enumerate(rows):
                row_num = i + 4
                fill = alt_fill if i % 2 == 1 else None
                values = [date_lbl, proj, hrs, desc]
                aligns = [center, left, right, left]
                ws.append(values)
                for col, (val, aln) in enumerate(zip(values, aligns), start=1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.alignment = aln
                    cell.border = cell_border
                    if fill:
                        cell.fill = fill
                    if col == 3:  # Hours column — format as number
                        cell.number_format = "0.00"
                ws.row_dimensions[row_num].height = 18

            # ── Total row ────────────────────────────────────────────
            if rows:
                ws.append([])
                total_row = len(rows) + 5
                ws.cell(total_row, 1).value = "Total"
                ws.cell(total_row, 1).font = total_lbl_font
                ws.cell(total_row, 1).fill = total_fill
                ws.cell(total_row, 1).alignment = center
                ws.cell(total_row, 1).border = cell_border
                ws.cell(total_row, 2).fill = total_fill
                ws.cell(total_row, 2).border = cell_border
                ws.cell(total_row, 3).value = total_hours
                ws.cell(total_row, 3).font = total_font
                ws.cell(total_row, 3).fill = total_fill
                ws.cell(total_row, 3).alignment = right
                ws.cell(total_row, 3).border = cell_border
                ws.cell(total_row, 3).number_format = "0.00"
                ws.cell(total_row, 4).fill = total_fill
                ws.cell(total_row, 4).border = cell_border
                ws.row_dimensions[total_row].height = 22

            # ── Column widths ─────────────────────────────────────────
            ws.column_dimensions["A"].width = 16
            ws.column_dimensions["B"].width = 28
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 52

            # Freeze panes below header
            ws.freeze_panes = "A4"

            wb.save(file_path)
            self.exportDone.emit(f"Exported to {os.path.basename(file_path)}", True)

        except Exception as e:
            self.exportDone.emit(f"Export failed: {e}", False)

    @Slot(str, str)
    def exportMonthlyReportCsv(self, year_month, file_path):
        """Export a monthly report to a CSV file — same rows as exportMonthlyReport,
        without the Excel formatting."""
        try:
            file_path = _url_to_path(file_path)
            if not file_path.endswith(".csv"):
                file_path += ".csv"

            year, month = int(year_month[:4]), int(year_month[5:7])
            month_name, rows, total_hours = self._collect_monthly_rows(year, month)

            with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow([f"Monthly Report — {month_name}"])
                writer.writerow([])
                writer.writerow(["Date", "Project", "Hours", "Description"])
                for date_lbl, proj, hrs, desc in rows:
                    writer.writerow([date_lbl, proj, f"{hrs:.2f}", desc])
                if rows:
                    writer.writerow([])
                    writer.writerow(["Total", "", f"{total_hours:.2f}", ""])

            self.exportDone.emit(f"Exported to {os.path.basename(file_path)}", True)

        except Exception as e:
            self.exportDone.emit(f"Export failed: {e}", False)

    @Slot(str)
    def exportJson(self, file_path: str):
        try:
            file_path = _url_to_path(file_path)
            if not file_path.endswith(".json"):
                file_path += ".json"
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(self._data, f, indent=2, ensure_ascii=False)
            self.jsonTransferDone.emit(f"Exported to {os.path.basename(file_path)}", True)
        except Exception as e:
            self.jsonTransferDone.emit(f"Export failed: {e}", False)

    @Slot(str)
    def importJson(self, file_path: str):
        try:
            file_path = _url_to_path(file_path)
            with open(file_path, "r", encoding="utf-8") as f:
                new_data = json.load(f)
            if "dailyLogs" not in new_data:
                self.jsonTransferDone.emit("Import failed: not a valid tracker data file", False)
                return
            snapshot_before_overwrite("pre-import")
            # Migrate legacy string-format projects
            new_data["projects"] = [
                p if isinstance(p, dict)
                else {"name": p, "billingCode": "", "billable": True}
                for p in new_data.get("projects", [])
            ]
            new_data.setdefault("tasks", [])
            self._data = new_data
            save_data(self._data)
            self._project_model.refresh()
            self._history_model.refresh()
            self._task_model.refresh()
            self.completedTasksChanged.emit()
            self.hasTodayLogsChanged.emit()
            self.summaryChanged.emit()
            self.jsonTransferDone.emit(f"Imported {os.path.basename(file_path)}", True)
        except Exception as e:
            self.jsonTransferDone.emit(f"Import failed: {e}", False)

    @Slot(result="QVariantList")
    def getAvailableBackups(self):
        """List every backup on disk (rolling, daily, pre-import, pre-restore),
        newest first by actual file modification time."""
        rows = []
        if os.path.exists(BAK_FILE):
            rows.append((os.path.getmtime(BAK_FILE), {
                "path": BAK_FILE,
                "label": "Most recent save",
                "kind": "rolling",
            }))
        for snap_date, path in _daily_backups():
            rows.append((os.path.getmtime(path), {
                "path": path,
                "label": snap_date.strftime("%a, %b %d, %Y"),
                "kind": "daily",
            }))
        for stamp, path in _named_backups("pre-import"):
            rows.append((os.path.getmtime(path), {
                "path": path,
                "label": "Before import — " + stamp.strftime("%b %d, %Y %I:%M %p"),
                "kind": "pre-import",
            }))
        for stamp, path in _named_backups("pre-restore"):
            rows.append((os.path.getmtime(path), {
                "path": path,
                "label": "Before restore — " + stamp.strftime("%b %d, %Y %I:%M %p"),
                "kind": "pre-restore",
            }))
        rows.sort(key=lambda r: r[0], reverse=True)
        return [r[1] for r in rows]

    @Slot(str)
    def restoreBackup(self, path: str):
        data = _try_load(path)
        if data is None:
            self.jsonTransferDone.emit("Restore failed: backup file is missing or invalid", False)
            return
        # Safety net: snapshot the current state before replacing it, so
        # restoring is itself undoable (same reasoning as the pre-import snapshot).
        snapshot_before_overwrite("pre-restore")
        if self._active_project:
            self.stopTimer()
        data["projects"] = [
            p if isinstance(p, dict)
            else {"name": p, "billingCode": "", "billable": True}
            for p in data.get("projects", [])
        ]
        data.setdefault("tasks", [])
        self._data = data
        save_data(self._data)
        self._project_model.refresh()
        self._history_model.refresh()
        self._task_model.refresh()
        self.completedTasksChanged.emit()
        self.hasTodayLogsChanged.emit()
        self.summaryChanged.emit()
        self.jsonTransferDone.emit("Restored from backup", True)

    # ── Internal ──

    def _save_active_session(self):
        self._data["activeSession"] = {
            "project": self._active_project,
            "startTime": self._session_start,
            "task": self._active_task_id,
        }
        save_data(self._data)

    def _clear_active_session(self):
        self._data.pop("activeSession", None)
        save_data(self._data)

    def _restore_active_session(self):
        session = self._data.get("activeSession")
        if not session:
            return
        project = session.get("project")
        start_time = session.get("startTime")
        if not project or not start_time:
            return
        # Verify the project still exists and is not archived
        known = [
            _proj_name(p) for p in self._data["projects"]
            if not (isinstance(p, dict) and p.get("archived", False))
        ]
        if project not in known:
            self._data.pop("activeSession", None)
            save_data(self._data)
            return
        self._active_project = project
        self._session_start = start_time
        self._last_checkin = time.time()
        self._elapsed = int(time.time() - start_time)
        task_id = session.get("task")
        task = self._task_by_id(task_id) if task_id else None
        self._active_task_id = task_id if task and not task.get("done") else None

    @Slot()
    def saveAndStop(self):
        """Save the active session time to disk without clearing session state.
        Called on app exit so in-progress time is not lost.
        """
        if self._active_project and self._session_start:
            self._log_time(int(time.time() - self._session_start))
            # Update the saved session start to now so the logged block
            # is not double-counted if the app is reopened and resumed.
            self._session_start = time.time()
            self._save_active_session()

    def _log_time(self, secs):
        if not self._active_project or secs <= 0:
            return
        session_start_ts = self._session_start if self._session_start else (time.time() - secs)
        start_dt = datetime.fromtimestamp(session_start_ts)
        end_dt = datetime.fromtimestamp(session_start_ts + secs)

        start_min = start_dt.hour * 60 + start_dt.minute
        # Clamp to same calendar day
        if end_dt.date() != start_dt.date():
            end_min = 23 * 60 + 59
        else:
            end_min = end_dt.hour * 60 + end_dt.minute

        day_key = start_dt.date().isoformat()
        if day_key not in self._data["dailyLogs"]:
            self._data["dailyLogs"][day_key] = {}
        if self._active_project not in self._data["dailyLogs"][day_key]:
            self._data["dailyLogs"][day_key][self._active_project] = {
                "seconds": 0,
                "sessions": [],
                "description": "",
            }
        entry = self._data["dailyLogs"][day_key][self._active_project]
        if "sessions" not in entry:
            entry["sessions"] = []
        entry["sessions"].append({"start": start_min, "end": end_min, "task": self._active_task_id})
        entry["seconds"] = sum((s["end"] - s["start"]) * 60 for s in entry["sessions"])
        if self._active_task_id:
            self._autofill_description(day_key, self._active_project, self._active_task_id)
        save_data(self._data)

    def _autofill_description(self, day_key, project, task_id):
        """Append the task's title to the day/project description, once, without
        touching any text already there (manual edits are never overwritten)."""
        task = self._task_by_id(task_id)
        if not task:
            return
        entry = self._data["dailyLogs"][day_key][project]
        lines = [l for l in entry.get("description", "").split("\n") if l.strip()]
        if task["title"] not in lines:
            lines.append(task["title"])
            entry["description"] = "\n".join(lines)

    def _task_by_id(self, task_id):
        for t in self._data.get("tasks", []):
            if t["id"] == task_id:
                return t
        return None

    def _task_seconds(self, task_id):
        total = 0
        for log in self._data.get("dailyLogs", {}).values():
            for info in log.values():
                for s in info.get("sessions", []):
                    if s.get("task") == task_id:
                        total += (s["end"] - s["start"]) * 60
        if self._active_task_id == task_id and self._session_start:
            total += int(time.time() - self._session_start)
        return total

    def _get_today_total(self, proj):
        today = date.today().isoformat()
        base = (
            self._data.get("dailyLogs", {})
            .get(today, {})
            .get(proj, {})
            .get("seconds", 0)
        )
        if self._active_project == proj and self._session_start:
            base += int(time.time() - self._session_start)
        return base

    def _tick(self):
        if self._active_project and self._session_start:
            self._elapsed = int(time.time() - self._session_start)
            self.elapsedChanged.emit()
            self.elapsedTextChanged.emit()
            self.summaryChanged.emit()
            # Refresh models to update "today" times for active project
            self._project_model.refresh()
            self._history_model.refresh()
            self._task_model.refresh_active_time()

        # Check-in
        if self._active_project and self._last_checkin:
            # Auto-stop if the check-in dialog has been unanswered for too long
            if self._checkin_shown_at and time.time() - self._checkin_shown_at >= self.INACTIVITY_TIMEOUT:
                self.stopTimer()
                return
            if time.time() - self._last_checkin >= self._data["checkInInterval"] * 60:
                self._last_checkin = time.time()
                self._checkin_shown_at = time.time()
                self.showCheckIn.emit()

        # EOD prompt
        if not self._eod_dismissed:
            now = datetime.now()
            today = date.today().isoformat()
            log = self._data.get("dailyLogs", {}).get(today, {})
            if now.hour >= 18 and log:
                has_time = any(v["seconds"] > 0 for v in log.values())
                all_descs = all(v.get("description") for v in log.values())
                if has_time and not all_descs:
                    self._eod_dismissed = True
                    self.openEodDialog()

    def _get_report_date_keys(self):
        today = date.today()
        if self._report_period == "day":
            target = today + timedelta(days=self._report_offset)
            return [target.isoformat()]
        elif self._report_period == "week":
            start = today - timedelta(days=today.weekday()) + timedelta(weeks=self._report_offset)
            return [(start + timedelta(days=i)).isoformat() for i in range(7)]
        elif self._report_period == "month":
            month = today.month + self._report_offset
            year = today.year
            while month <= 0:
                month += 12
                year -= 1
            while month > 12:
                month -= 12
                year += 1
            days_in_month = calendar.monthrange(year, month)[1]
            return [date(year, month, d + 1).isoformat() for d in range(days_in_month)]
        return []

    def _get_report_label(self):
        today = date.today()
        if self._report_period == "day":
            target = today + timedelta(days=self._report_offset)
            if self._report_offset == 0:
                return "Today"
            elif self._report_offset == -1:
                return "Yesterday"
            return target.strftime("%a, %b %d, %Y")
        elif self._report_period == "week":
            start = today - timedelta(days=today.weekday()) + timedelta(weeks=self._report_offset)
            end = start + timedelta(days=6)
            if self._report_offset == 0:
                return f"This Week ({start.strftime('%b %d')} \u2013 {end.strftime('%b %d')})"
            return f"{start.strftime('%b %d')} \u2013 {end.strftime('%b %d, %Y')}"
        elif self._report_period == "month":
            month = today.month + self._report_offset
            year = today.year
            while month <= 0:
                month += 12
                year -= 1
            while month > 12:
                month -= 12
                year += 1
            target = date(year, month, 1)
            if self._report_offset == 0:
                return f"This Month ({target.strftime('%B %Y')})"
            return target.strftime("%B %Y")
        return ""

    def _load_report(self):
        keys = self._get_report_date_keys()
        self._report_model.load(keys)
        self.reportLabelChanged.emit()
        self.reportTotalChanged.emit()
        self.reportTotalSecondsChanged.emit()

    @Slot(str, str, float, float, result='QVariantMap')
    def calculateUtilization(self, start_date: str, end_date: str, pto_hours: float, holiday_hours: float):
        """Calculate utilization rates for a date range.

        Standard working hours are fixed at 8 h/day (Mon–Fri).
        pto_hours and holiday_hours are deducted from standard hours for the adjusted rate.

        Returns a dict with:
          billableHours, totalHours, workingDays, ptoHours, holidayHours,
          standardHours, adjustedHours, rate1, rate2, rate3
        where rate values are percentages (0–100) or -1 when denominator is 0.
        """
        try:
            start = date.fromisoformat(start_date)
            end = date.fromisoformat(end_date)
        except Exception:
            return {"error": "Invalid date format. Use YYYY-MM-DD."}

        if end < start:
            return {"error": "End date must be on or after start date."}

        # Build set of billable project names (includes archived)
        billable_projects = set()
        for p in self._data["projects"]:
            if isinstance(p, dict):
                if p.get("billable", True):
                    billable_projects.add(p["name"])
            else:
                billable_projects.add(p)

        # Sum billable and total seconds; count Mon–Fri working days
        billable_secs = 0
        total_secs = 0
        working_days = 0
        cur = start
        while cur <= end:
            dk = cur.isoformat()
            log = self._data["dailyLogs"].get(dk, {})
            for proj, info in log.items():
                secs = info.get("seconds", 0)
                total_secs += secs
                if proj in billable_projects:
                    billable_secs += secs
            if cur.weekday() < 5:  # Monday–Friday
                working_days += 1
            cur += timedelta(days=1)

        billable_hours = billable_secs / 3600
        total_hours = total_secs / 3600
        standard_hours = working_days * 8
        adjusted_hours = standard_hours - pto_hours - holiday_hours

        def pct(num, den):
            if den <= 0:
                return -1.0
            return round(num / den * 100, 1)

        return {
            "billableHours": round(billable_hours, 2),
            "totalHours": round(total_hours, 2),
            "workingDays": working_days,
            "ptoHours": pto_hours,
            "holidayHours": holiday_hours,
            "standardHours": standard_hours,
            "adjustedHours": adjusted_hours,
            "rate1": pct(billable_hours, total_hours),
            "rate2": pct(billable_hours, standard_hours),
            "rate3": pct(billable_hours, adjusted_hours),
        }
